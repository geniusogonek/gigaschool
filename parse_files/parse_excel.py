import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime

def extract_date_from_sheet(sheet_df):
    """
    Ищет дату в любом месте листа.
    Работает с форматами dd.mm.yyyy, dd/mm/yyyy, yyyy-mm-dd.
    """
    date_pattern = r"(\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b)"
    
    for row in sheet_df.astype(str).values:
        for cell in row:
            match = re.search(date_pattern, cell)
            if match:
                try:
                    return pd.to_datetime(match.group(0), dayfirst=True).date()
                except:
                    pass
    return None


def parse_schedule(file_path, class_name):
    wb = load_workbook(file_path, data_only=True)
    result = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.DataFrame(ws.values)

        # --- 1. Поиск даты на листе ---
        date_on_sheet = extract_date_from_sheet(df)
        
        # Если лист без даты — пропускаем
        if not date_on_sheet:
            continue

        # --- 2. Поиск колонки класса ---
        header_row = df.iloc[0].astype(str).str.lower()
        class_col_index = None
        
        for idx, val in enumerate(header_row):
            if class_name.lower() in val:
                class_col_index = idx
                break
        
        if class_col_index is None:
            continue  # В этом листе нет нужного класса

        # --- 3. Извлечение урок → кабинет ---
        lessons = []
        for row in df.iloc[1:].values:
            lesson_info = row[class_col_index]
            if lesson_info and str(lesson_info).strip() not in ["", "nan"]:
                lessons.append(str(lesson_info))

        result.append({
            "sheet": sheet,
            "date": date_on_sheet,
            "lessons": lessons
        })

    return result


# ------------------
# Пример использования
# ------------------

file_path = "schedule.xlsx"
class_name = "7А"

parsed = parse_schedule(file_path, class_name)

for entry in parsed:
    print("Лист:", entry["sheet"])
    print("Дата:", entry["date"])
    print("Уроки и кабинеты:")
    for lesson in entry["lessons"]:
        print(" •", lesson)
    print()
